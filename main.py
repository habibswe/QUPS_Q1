import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime

# Define the path to your Excel file
excel_path = r"C:\Users\habib\PycharmProjects\Q1Test\Excel.xlsx"

# Define a function to get Google search suggestions for a given keyword
def get_google_suggestions(keyword):
    driver = webdriver.Chrome()
    driver.get("https://www.google.com")
    search_box = driver.find_element(By.NAME, "q")
    search_box.clear()
    search_box.send_keys(keyword)

    wait = WebDriverWait(driver, 5)
    wait.until(EC.presence_of_element_located((By.XPATH, "//ul[@role='listbox']/li[@role='presentation']")))

    Google_suggestions = driver.find_elements(By.XPATH, "//ul[@role='listbox']/li[@role='presentation']")
    google_suggestion_texts = [suggestion.text for suggestion in Google_suggestions]

    driver.quit()

    if google_suggestion_texts:
        longest_suggestion = max(google_suggestion_texts, key=len)
        shortest_suggestion = min(google_suggestion_texts, key=len)
        return longest_suggestion, shortest_suggestion
    else:
        return None,


assignment = openpyxl.load_workbook(excel_path)

# Get the current day
current_day = datetime.now().strftime("%A")

if current_day in assignment.sheetnames:
    worksheet = assignment[current_day]

    longest_suggestions = []
    shortest_suggestions = []

    # Iterate through the rows and get suggestions
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2, max_col=2):
        keyword = row[0].value
        if keyword and not keyword.isspace():
            long_suggestion, short_suggestion = get_google_suggestions(keyword)
            longest_suggestions.append(long_suggestion)
            shortest_suggestions.append(short_suggestion)

    for idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=3, max_col=4)):
        row[0].value = longest_suggestions[idx]
        row[1].value = shortest_suggestions[idx]

    # Save the updated Excel file
    assignment.save(excel_path)

assignment.close()
